from openpyxl import *
from openpyxl import load_workbook
import tkinter as tk
import warnings
import os

def editMacroFile(path, entries):
    try:
        # TESTING
        # path = f'C:/Users/Reagan Toh/Documents/AvayaReportTesting/Avaya_Report latest.xlsm'

        # Directory for macro file
        filename = f"{path}/Avaya_Report latest.xlsm"
        warnings.simplefilter(action='ignore', category=UserWarning)
        wb = load_workbook(filename=filename, read_only=False, keep_vba=True)

        # Open workbook
        ws = wb.worksheets[0]


        # TESTING
        # ws["B1"] = f"C:/Users/{user}/Documents/AvayaReportTesting/"
        print(path)
        # Directory for report
        ws["B1"] = path + "/"

        # Assigning variables to each file needed for macro
        
        for entry in entries:
            if entry.startswith("Agent All"):
                agentAll = entry

            elif entry.startswith("Avaya Reports"):
                avayaReport = entry

            elif entry.startswith("Contact_Details"):
                contactDetails = entry

            elif entry.startswith("CRVReport"):
                CRVReport = entry

            elif entry.startswith("Topics by All hours"):
                allHours = entry

            elif entry.startswith("Topics by Days"):
                topicsByDays = entry

        # Editing macro file to run macro
        ws["B2"] = agentAll
        ws["B3"] = topicsByDays
        ws["B4"] = allHours
        ws["B5"] = contactDetails
        ws["B6"] = CRVReport
        ws["B7"] = avayaReport

        # Save workbook
        wb.save(filename=filename)
        wb.close()

        return filename

    # In the event that the user tries to run the program while trying to work inside the oneDrive directory,
    # Tkinter GUI will inform the user that the program has no permission to edit oneDrive files and will be asked
    # to work inside local drive instead.

    except PermissionError:
        # Tkinter GUI for invalid directory / no permission
        ws = tk.Tk()
        ws.geometry("300x200")
        ws.title("Error!")

        w = tk.Label(ws, text="No Permission to folder/file, try again on local folder", font="50")
        w.pack(side=tk.TOP, pady=20)

        tk.Button(ws, text="Ok", command=ws.quit).pack(side=tk.BOTTOM, pady=20)

        ws.mainloop()

        exit()

        


