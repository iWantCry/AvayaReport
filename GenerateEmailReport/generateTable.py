from openpyxl import load_workbook
from openpyxl import *
import tkinter as tk

from copyRange import copyRange
from pasteRange import pasteRange

def generateTable(path, day, monthNameAbbr):
    try:
        wb = load_workbook(path)
        ws = wb.active

        ws["A1"] = f"Avaya Call Volume {day-1} {monthNameAbbr}"

        # Remove unwanted columns
        ws.delete_cols(13,2)
        ws.delete_cols(14,2)
        ws.delete_cols(19)

        # Copy and paste top data
        selectedRange = copyRange(12,2,19,5,ws) #Change the 4 number values
        pasteRange(2,3,9,6,ws,selectedRange) #Change the 4 number values

        # Copy and paste first half of bottom data
        selectedRange = copyRange(12,11,19,16,ws) #Change the 4 number values
        pasteRange(2,9,9,14,ws,selectedRange) #Change the 4 number values

        # Copy and paste second half of bottom data
        selectedRange = copyRange(12,19,19,23,ws) #Change the 4 number values
        pasteRange(2,15,9,19,ws,selectedRange) #Change the 4 number values

        ws.delete_rows(20,2)

        wb.save(filename=path)
        wb.close()

    except PermissionError:
        print("Error! File opened")

        # Tkinter GUI if file is still opened
        ws = tk.Tk()
        ws.geometry("300x200")
        ws.title("Error!")

        w = tk.Label(ws, text="Excel Workbook is currently opened!\nClose Excel Workbook and try again!", font="50")
        w.pack(side=tk.TOP, pady=20)

        tk.Button(ws, text="Ok", command=ws.quit).pack(side=tk.BOTTOM, pady=20)

        ws.mainloop()

        exit()





